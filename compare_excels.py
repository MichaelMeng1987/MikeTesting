#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

CHANGE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare two .xlsx workbooks and create a diff workbook showing mismatches."
    )
    parser.add_argument("left_file", help="First Excel file to compare")
    parser.add_argument("right_file", help="Second Excel file to compare")
    parser.add_argument(
        "--output-file",
        default="diff.xlsx",
        help="Path to save the comparison report workbook (default: diff.xlsx)",
    )
    return parser.parse_args()


def normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def value_to_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def safe_sheet_title(title: str) -> str:
    normalized = title.replace("\\", "-").replace("/", "-").replace("?", "-").replace("*", "-")
    normalized = normalized.replace("[", "(").replace("]", ")").replace(":", "-")
    if len(normalized) > 31:
        return normalized[:28].rstrip() + "..."
    return normalized


def build_summary_sheet(workbook: Workbook, summary_rows: list[dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "SUMMARY"
    headers = ["Sheet Name", "Status", "Differences", "Notes"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
    for row in summary_rows:
        sheet.append([row["sheet_name"], row["status"], row["differences"], row["notes"]])
    for column_index in range(1, 5):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20


def build_diff_sheet(
    workbook: Workbook,
    sheet_name: str,
    left_ws: Any | None,
    right_ws: Any | None,
) -> int:
    report_name = safe_sheet_title(f"Diff: {sheet_name}")
    report_sheet = workbook.create_sheet(report_name)
    headers = ["Row", "Column", "Left value", "Right value", "Status"]
    report_sheet.append(headers)
    for cell in report_sheet[1]:
        cell.font = HEADER_FONT

    if left_ws is None:
        report_sheet.append(["", "", "<Sheet missing in left file>", "", "LEFT_MISSING"])
        return 0
    if right_ws is None:
        report_sheet.append(["", "", "", "<Sheet missing in right file>", "RIGHT_MISSING"])
        return 0

    row_count = max(left_ws.max_row or 0, right_ws.max_row or 0)
    col_count = max(left_ws.max_column or 0, right_ws.max_column or 0)
    differences = 0

    for row in range(1, row_count + 1):
        for col in range(1, col_count + 1):
            left_value = normalize_value(left_ws.cell(row=row, column=col).value)
            right_value = normalize_value(right_ws.cell(row=row, column=col).value)
            if left_value != right_value:
                differences += 1
                cell_ref = get_column_letter(col)
                report_sheet.append(
                    [
                        row,
                        cell_ref,
                        value_to_display(left_value),
                        value_to_display(right_value),
                        "CHANGED",
                    ]
                )
                for changed_cell in report_sheet[row + 1][2:4]:
                    changed_cell.fill = CHANGE_FILL

    if differences == 0:
        report_sheet.append(["", "", "No differences found", "", "OK"])
    else:
        report_sheet.column_dimensions[get_column_letter(3)].width = 40
        report_sheet.column_dimensions[get_column_letter(4)].width = 40

    return differences


def compare_workbooks(left_path: Path, right_path: Path, output_path: Path) -> None:
    left_wb = load_workbook(filename=left_path, data_only=True)
    right_wb = load_workbook(filename=right_path, data_only=True)
    report_wb = Workbook()

    summary_rows: list[dict[str, Any]] = []
    left_sheets = set(left_wb.sheetnames)
    right_sheets = set(right_wb.sheetnames)
    all_sheets = sorted(left_sheets.union(right_sheets))

    for sheet_name in all_sheets:
        left_ws = left_wb[sheet_name] if sheet_name in left_sheets else None
        right_ws = right_wb[sheet_name] if sheet_name in right_sheets else None
        differences = build_diff_sheet(report_wb, sheet_name, left_ws, right_ws)
        status = "OK" if differences == 0 else "DIFFERENCES"
        if left_ws is None:
            status = "MISSING_IN_LEFT"
        elif right_ws is None:
            status = "MISSING_IN_RIGHT"
        summary_rows.append(
            {
                "sheet_name": sheet_name,
                "status": status,
                "differences": differences,
                "notes": "Sheet exists only in one file" if left_ws is None or right_ws is None else "",
            }
        )

    build_summary_sheet(report_wb, summary_rows)
    report_wb.save(output_path)
    print(f"Comparison complete. Diff workbook saved to: {output_path}")


def main() -> None:
    args = parse_args()
    left_path = Path(args.left_file)
    right_path = Path(args.right_file)
    output_path = Path(args.output_file)

    if not left_path.exists():
        raise FileNotFoundError(f"Left file not found: {left_path}")
    if not right_path.exists():
        raise FileNotFoundError(f"Right file not found: {right_path}")

    compare_workbooks(left_path, right_path, output_path)


if __name__ == "__main__":
    main()
